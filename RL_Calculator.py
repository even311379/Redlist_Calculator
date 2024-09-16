from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# wb = load_workbook(filename = 'Target.xlsx')
# sheet = wb['Assess_datatable']

def FormatData(FileName):
    ## store in this format
    ## {spe:{a:[], b:[], cd:[], e:[], rc:[]}}
    out = dict()
    sheet = load_workbook(FileName)['Assess_datatable']
    for c in range(5, sheet.max_column + 1):
        spe = sheet.cell(2, c).value
        out[spe] = dict(a=[], b=[], cd=[], e=[], rc=[])
        for r in range(5, 18):
            out[spe]['a'].append(sheet.cell(r, c).value)
        for r in range(18, 56):
            out[spe]['b'].append(sheet.cell(r, c).value)
        for r in range(56, 66):
            out[spe]['cd'].append(sheet.cell(r, c).value)
        out[spe]['e'].append(sheet.cell(66, c).value)
        for r in range(67, 73):
            out[spe]['rc'].append(sheet.cell(r, c).value)
    return out

## result = ["Category", "Criteria"]
## ex: ["DD", ""] ["LC", ""] ["NT", "B1a"] ["VU", "A2bcd"]

def GetA(a_data):
    decline_before = a_data[3]
    decline_after = a_data[8]
    if not decline_before:
        decline_before = 0
    if not decline_after:
        decline_after = 0
    decline = decline_before + decline_after
    if decline == 0:
        return ["DD", ""]
    decline_time_unit = a_data[5]

    ## TODO: there is no way to get decline time in FUTURE, only past in included in source data
    decline_time_length = a_data[4]

    if (decline_time_unit == 'Y' and decline_time_length >= 10) or \
    (decline_time_unit == 'G' and decline_time_length > 3) :
        out = []
        has_ceased = a_data[10] == "B" 
        if (decline <= -90 and has_ceased) or decline <= -70:
            out.append("CR")
        elif (decline <= 70 and has_ceased) or decline <= -50:
            out.append("EN")
        elif (decline <= 50 and has_ceased) or decline <= -30:
            out.append("VU")
        elif (decline <= 30 and has_ceased) or decline <= -20:
            out.append("NT")
        else:
            return ["LC", ""]
        if has_ceased:
            out.append("A1")
        else:
            if decline_before != 0 and decline_after != 0:
                out.append("A4")
            elif decline_after != 0:
                out.append("A3")
            else:
                out.append("A2")
        decline_source = a_data[7]
        if decline_source == "A":
            out[-1] += 'a'
        decline_data_type = a_data[6]
        if decline_data_type:
            if 'B' in decline_data_type:
                out[-1] += 'b'
            if 'C' in decline_data_type:
                out[-1] += 'c'
        threat_type = a_data[9]
        if 'B' in threat_type:
            out[-1] += 'd'
        for t in "CDEFGH":
            if t in threat_type:
                out[-1] += 'e'
                break
        return out
    return ["DD", ""]


def GetB(b_data):
    AOO = b_data[0] 
    EOO = b_data[1]
    if not AOO and not EOO:
        return ["DD", ""]
    if not AOO: # prevent empty error
        AOO = 99999
    if not EOO:
        EOO = 999999
    if type(EOO) == str:
        EOO = 999999    

    ## location
    locations = b_data[14]
    if not locations:
        locations = 999

    ## check decline
    declines = [False] * 5
    if b_data[17] == 'A': # continue decline in AOO/EOO
        if 'A' in b_data[18]:
            declines[1] = True
        for c in "BCD":
            if c in b_data[18]:
                declines[0] = True
                break

    ## TODO: there is no way to check decline in habitat!!!
        ## declines[2] (iii) will always be False...    


    if b_data[23] == 'A': # decline in population
        declines[3] = True
    if b_data[26] == 'A': # continue decline in mature indivisuals
        declines[4] = True

    ## check flutuation
    flutuations = [False] * 4
    if b_data[29] == 'Y': # AOO/EOO
        if 'A' in b_data[30]:
            flutuations[0] = True
        for c in "BCD":
            if c in b_data[30]:
                flutuations[1] = True
                break
    if b_data[33] == 'Y': # population
        flutuations[2] = True
    if b_data[35] == 'Y': # mature individuals
        flutuations[3] = True
    ## check false
    cons = (locations <= 10) + (sum(declines) > 0) + (sum(flutuations) > 0)
    if cons == 0:
        return ["LC", ""]
    if AOO >= 2000 and EOO >= 20000:
        return ["LC", ""]

    out = []
    if cons == 1:
        out.append("NT")
    ## TODO: include locations as level judgement?


    elif AOO < 10 or EOO < 100:
        out.append("CR")
    elif AOO < 500 or EOO < 5000:
        out.append("EN")
    else:
        out.append("VU")
    
    rs = "" ## reason string
    if locations <= 10:
        rs += 'a'
    if sum(declines) > 0:
        rs += 'b('
        for b, notation in zip(declines, ["i", "ii", "iii", "iv", "v"]):
            if b:
                rs += notation + ','
        rs = rs[:-1]
        rs += ')'
    if sum(flutuations) > 0:
        rs += 'c('
        for b, notation in zip(flutuations, ["i", "ii", "iii", "iv"]):
            if b:
                rs += notation + ','
        rs = rs[:-1]
        rs += ')'
    if AOO < 2000 and EOO < 20000:
        out.append(f"B1{rs}+B2{rs}")
    elif EOO < 20000:
        out.append(f"B1{rs}")
    else:
        out.append(f"B2{rs}")
    return out


def GetC(data):
    a_data = data['a']
    b_data = data['b']
    cd_data = data['cd']
    if not cd_data[0]:
        return ["DD", ""]
    if type(cd_data[0]) == str:
        print(f"Error! should be number, {cd_data[0]}")
        return ["DD", ""]
    if cd_data[0] >= 20000:
        return ["LC", ""]

    # check C1
    # TODO: the C1 decline description is not robust enough to make judgement...
    # use ratio instead... somehow, is not precise ...
    C1_passed = False
    decline_before = a_data[3]
    decline_after = a_data[8]
    decline_time_unit = a_data[5]
    decline_time_length = a_data[4]  # TODO: same FUTURE issue as in A
    if not decline_before:
        decline_before = 0
    if not decline_after:
        decline_after = 0
    decline = decline_before + decline_after
    if decline < 0:
        # translation is wrong in C1 in taiwan redlist of freshwater fish
        # WRONG... it's definitely irrelavant to treat it in linear way...
        if decline_time_unit == 'Y':
            C1_passed = decline / decline_time_length <= -1
        elif decline_time_unit == 'G':
            C1_passed = decline / decline_time_length <= -3.33

    # check C2
    C2_passed = False
    # TODO: C2a is impossible to check in your data table!!! Skip
    # C2a description is so vague... (at least I don't know what it is about)
    # only use C2b for now
    if b_data[35] == 'Y':
        C2_passed = True

    # eval criteria ## TODO: (skip C1 and C2a to judge criteria)
    if (C1_passed + C2_passed) == 0:
        return ["LC", ""]
    out = []
    n_individuals = cd_data[0]
    if n_individuals < 250:
        out.append("CR")
    elif n_individuals < 2500:
        out.append("EN")
    elif n_individuals < 10000:
        out.append("VU")
    else:
        out.append("NT")

    ## reason string
    rs = ""

    ## TODO: reason for C2a is skipped...

    if b_data[35] == 'Y':
        rs += 'b'

    if C1_passed and C2_passed:
        out.append(f"C1+C2{rs}")
    elif C1_passed:
        out.append("C1")
    else:
        out.append("C2{rs}")

    return out


def GetD(data):
    b_data = data['b']    
    cd_data = data['cd']

    out = []
    if cd_data[0]:
        if cd_data[0] <= 50:
            out.append("CR")
            out.append("D") ## TODO: return D or D1?
        elif cd_data[0] <= 250:
            out.append("EN")
            out.append("D")
        elif cd_data[0] <= 1000:
            out.append("VU")
            out.append("D1")
        elif cd_data[0] <= 2500:
            out.append("NT") 
            out.append("D1")
        if cd_data[0] <= 2500:
            return out
    else:
        return ["DD", ""]

    AOO = b_data[0]
    if not AOO: 
        AOO = 99999
    locations = b_data[14]
    if not locations:
        locations = 999999
    if AOO <= 20 or locations <= 5:
        return ["VU", "D2"]
    if AOO <= 50 or locations <= 10:
        return ["NT", "D2"]
    return ["LC", ""]


def GetE(e_data):
    if not e_data[0]:
        return ["DD", ""]

    ## TODO: ignore it, the data in data table is just wrong...
        # it's definitely not linear...
    return ["LC", ""]

CriteriaValueMap = {"CR":0, "EN":1, "VU":2, "NT":3, "LC":4, "DD":5}
CriteriaValueColorMap = {"CR":"FF3737", "EN":"FF6C60", "VU":"FFC080", "NT":"FF8C69"}

def ValueToCriteria(v):
    if v == 0:
        return "CR"
    if v == 1:
        return "EN"
    if v == 2:
        return "VU"
    if v == 3:
        return "NT"
    return "LC"

def Assess(data):
    out_file = "assess_result.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "results"
    spe_list = list(data.keys())
    ws.cell(1,1,value="Species")
    ws.cell(1,2,value="A")
    ws.cell(1,3,value="B")
    ws.cell(1,4,value="C")
    ws.cell(1,5,value="D")
    ws.cell(1,6,value="E")
    ws.cell(1,7,value="Region Adjustment")
    ws.cell(1,8,value="Final")
    for c, s in enumerate(spe_list):
        results = [
            GetA(data[s]['a']),
            GetB(data[s]['b']),
            GetC(data[s]),
            GetD(data[s]),
            GetE(data[s]['e'])
        ]
        ws.cell(c+2, 1, value=s)
        r_adj = data[s]['rc'][4]
        ws.cell(c+2, 7, value=r_adj*-1)
        LowestValue = 5
        LowestResults = []
        for r, result in enumerate(results):
            ws.cell(c+2, r+2, value=f"{result[0]} {result[1]}")

            if result[0] in CriteriaValueColorMap:
                color = CriteriaValueColorMap[result[0]]
                ws.cell(c+2, r+2).fill = PatternFill(color, color, fill_type="solid")

            if CriteriaValueMap[result[0]] < LowestValue:
                LowestValue = CriteriaValueMap[result[0]]
        if LowestValue == 5:
            ws.cell(c+2, 8, value="DD")
            continue
        for result in results:
            if CriteriaValueMap[result[0]] == LowestValue:
                LowestResults.append(result[1])
        rs = ", ".join(LowestResults)
        if LowestValue > 3:
            rs = ""
        if r_adj != 0:
            LowestValue += r_adj
            if LowestValue < 0:
                LowestValue = 0
            if LowestValue > 4:
                LowestValue = 4
            ws.cell(c+2, 8, value=f"{ValueToCriteria(LowestValue)}[{r_adj*-1}] {rs}")
        else:
            ws.cell(c+2, 8, value=f"{ValueToCriteria(LowestValue)} {rs}")

        final_criteria = ValueToCriteria(LowestValue)
        if final_criteria in CriteriaValueColorMap:
            color = CriteriaValueColorMap[final_criteria]
            ws.cell(c+2, 8).fill = PatternFill(color, color, fill_type="solid")


    wb.save(filename = out_file)


if __name__ == '__main__':
    data = FormatData("Target.xlsx")
    Assess(data)
    # print(data)
    # spe_list = list(data.keys())
    # for i,v in enumerate(data["Aphyocypris amnis Liao, Kullander & Lin, 2011"]['cd']):
    #     print(i, v)
    # if not data[spe_list[0]]['a'][4]:
    #     print('empty')
    # print(data["Aphyocypris amnis Liao, Kullander & Lin, 2011"]['cd'])
    # print(GetC(data["Aphyocypris amnis Liao, Kullander & Lin, 2011"]['cd']))
